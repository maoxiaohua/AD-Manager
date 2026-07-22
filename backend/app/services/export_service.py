import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models import Computer, User, ADGroup, GroupMembership


class ExportService:
    EXPORT_COLUMNS = {
        "computers": [
            "name", "distinguished_name", "ip_address", "operating_system",
            "os_version", "description", "status",
            "last_logon_timestamp", "created_at",
        ],
        "users": [
            "sam_account_name", "display_name", "email", "department",
            "distinguished_name", "site", "hostnames", "hostname_count", "created_at",
        ],
        "groups": [
            "name", "distinguished_name", "display_name", "group_type",
            "group_scope", "description", "email",
            "end_user_email", "jira_ticket", "created_at",
        ],
        "user-bindings": [
            "sam_account_name", "display_name", "department", "site",
            "hostname", "group_type", "group_description",
        ],
    }

    def __init__(self, db: Session):
        self.db = db

    def export_csv(self, entity_type: str, filters: dict | None = None, template: bool = False) -> io.StringIO:
        """Build a CSV in memory and return the StringIO buffer."""
        if entity_type not in self.EXPORT_COLUMNS:
            raise ValueError(f"Unknown entity type: {entity_type}")

        columns = self.EXPORT_COLUMNS[entity_type]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()

        if not template:
            rows = self._fetch_rows(entity_type, filters)
            for row_dict in rows:
                formatted = {}
                for k, v in row_dict.items():
                    if hasattr(v, "isoformat"):
                        formatted[k] = v.isoformat() if v else ""
                    else:
                        formatted[k] = str(v) if v is not None else ""
                writer.writerow(formatted)

        output.seek(0)
        return output

    def export_excel(self, entity_type: str, filters: dict | None = None, template: bool = False) -> io.BytesIO:
        """Build a styled .xlsx in memory and return the BytesIO buffer."""
        if entity_type not in self.EXPORT_COLUMNS:
            raise ValueError(f"Unknown entity type: {entity_type}")

        columns = self.EXPORT_COLUMNS[entity_type]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = entity_type.title()

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        if not template:
            rows = self._fetch_rows(entity_type, filters)
            for row_idx, row_dict in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = row_dict.get(col_name)
                    if hasattr(val, "isoformat"):
                        val = val.isoformat() if val else ""
                    ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_length:
                    max_length = cell_len
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _fetch_rows(self, entity_type: str, filters: dict | None = None) -> list[dict]:
        """Fetch and format rows from the database."""
        filters = filters or {}

        if entity_type == "computers":
            query = self.db.query(Computer)
            if filters.get("search"):
                s = f"%{filters['search']}%"
                query = query.filter(
                    (Computer.name.ilike(s))
                    | (Computer.ip_address.ilike(s))
                    | (Computer.description.ilike(s))
                    | (Computer.distinguished_name.ilike(s))
                )
            if filters.get("status"):
                query = query.filter(Computer.status == filters["status"])

            results = []
            for comp in query.order_by(Computer.name).all():
                results.append({
                    "name": comp.name,
                    "distinguished_name": comp.distinguished_name,
                    "ip_address": comp.ip_address or "",
                    "operating_system": comp.operating_system or "",
                    "os_version": comp.os_version or "",
                    "description": comp.description or "",
                    "status": comp.status,
                    "last_logon_timestamp": comp.last_logon_timestamp,
                    "created_at": comp.created_at,
                })
            return results

        elif entity_type == "users":
            query = self.db.query(User)
            if filters.get("search"):
                s = f"%{filters['search']}%"
                query = query.filter(
                    (User.sam_account_name.ilike(s))
                    | (User.display_name.ilike(s))
                    | (User.email.ilike(s))
                    | (User.department.ilike(s))
                )
            results = []
            for user in query.order_by(User.sam_account_name).all():
                # Resolve group memberships
                memberships = (
                    self.db.query(GroupMembership)
                    .filter(GroupMembership.user_id == user.id)
                    .all()
                )
                hostnames = []
                for m in memberships:
                    g = self.db.query(ADGroup).filter(ADGroup.id == m.group_id).first()
                    if g:
                        hostnames.append(g.name)

                # Parse site from DN
                from app.core.dn_parser import parse_dn
                dn_info = parse_dn(user.distinguished_name)

                results.append({
                    "sam_account_name": user.sam_account_name,
                    "display_name": user.display_name or "",
                    "email": user.email or "",
                    "department": user.department or "",
                    "distinguished_name": user.distinguished_name,
                    "site": dn_info["site"] or "",
                    "hostnames": ", ".join(hostnames),
                    "hostname_count": str(len(hostnames)),
                    "created_at": user.created_at,
                })
            return results

        elif entity_type == "groups":
            query = self.db.query(ADGroup)
            results = []
            for group in query.order_by(ADGroup.name).all():
                results.append({
                    "name": group.name,
                    "distinguished_name": group.distinguished_name,
                    "display_name": group.display_name or "",
                    "group_type": group.group_type,
                    "group_scope": group.group_scope,
                    "description": group.description or "",
                    "email": group.email or "",
                    "end_user_email": group.end_user_email or "",
                    "jira_ticket": group.jira_ticket or "",
                    "created_at": group.created_at,
                })
            return results

        elif entity_type == "user-bindings":
            from app.core.dn_parser import parse_dn
            memberships = self.db.query(GroupMembership).all()
            results = []
            for m in memberships:
                user = self.db.query(User).filter(User.id == m.user_id).first()
                group = self.db.query(ADGroup).filter(ADGroup.id == m.group_id).first()
                if not user or not group:
                    continue
                dn_info = parse_dn(user.distinguished_name)
                results.append({
                    "sam_account_name": user.sam_account_name,
                    "display_name": user.display_name or "",
                    "department": user.department or "",
                    "site": dn_info["site"] or "",
                    "hostname": group.name,
                    "group_type": group.group_type,
                    "group_description": group.description or "",
                })
            return sorted(results, key=lambda r: r["sam_account_name"])

        return []
