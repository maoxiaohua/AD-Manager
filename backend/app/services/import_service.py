import csv
import io
from datetime import datetime, timezone
from sqlalchemy.orm import Session
from app.models import SyncLog, Computer, User, ADGroup


class ImportService:
    REQUIRED_COLUMNS = {
        "computers": ["name", "distinguished_name"],
        "users": ["sam_account_name", "distinguished_name"],
        "groups": ["name", "distinguished_name"],
    }
    OPTIONAL_COLUMNS = {
        "computers": ["ip_address", "operating_system", "os_version", "description", "status"],
        "users": ["display_name", "email", "department"],
        "groups": ["display_name", "group_type", "group_scope", "description", "email"],
    }

    def __init__(self, db: Session):
        self.db = db

    def import_csv(self, file_content: bytes, entity_type: str) -> SyncLog:
        reader = csv.DictReader(io.StringIO(file_content.decode("utf-8-sig")))
        self._validate_columns(reader.fieldnames or [], entity_type)
        return self._process_rows(reader, entity_type, "import")

    def import_excel(self, file_content: bytes, entity_type: str) -> SyncLog:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = [str(cell) if cell else "" for cell in next(rows_iter)]
        self._validate_columns(header, entity_type)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(header, [str(v) if v is not None else "" for v in row]))
            rows.append(row_dict)
        wb.close()
        return self._process_rows(rows, entity_type, "import")

    def _process_rows(self, rows, entity_type: str, sync_type: str) -> SyncLog:
        log = SyncLog(
            sync_type=sync_type,
            status="running",
            started_at=datetime.now(timezone.utc),
        )
        self.db.add(log)
        self.db.commit()

        count = 0
        try:
            for row in rows:
                if not any(v for v in row.values() if v):
                    continue
                if entity_type == "computers":
                    self._import_computer_row(row)
                elif entity_type == "users":
                    self._import_user_row(row)
                elif entity_type == "groups":
                    self._import_group_row(row)
                count += 1

                if count % 100 == 0:
                    self.db.commit()

            self.db.commit()
            log.status = "success"
            log.records_processed = count
        except Exception as e:
            self.db.rollback()
            log.status = "failed"
            log.error_message = f"Row {count + 1}: {str(e)}"

        log.completed_at = datetime.now(timezone.utc)
        self.db.commit()
        return log

    def _import_computer_row(self, row: dict):
        dn = row.get("distinguished_name", "").strip()
        if not dn:
            raise ValueError(f"Missing distinguished_name for computer: {row.get('name', 'unknown')}")

        existing = self.db.query(Computer).filter(Computer.distinguished_name == dn).first()
        if existing:
            for key in ["name", "ip_address", "operating_system", "os_version", "description", "status"]:
                if key in row and row[key]:
                    setattr(existing, key, row[key].strip())
        else:
            computer = Computer(
                name=row.get("name", "").strip(),
                distinguished_name=dn,
                ip_address=row.get("ip_address", "").strip() or None,
                operating_system=row.get("operating_system", "").strip() or None,
                os_version=row.get("os_version", "").strip() or None,
                description=row.get("description", "").strip() or None,
                status=row.get("status", "active").strip(),
            )
            self.db.add(computer)

    def _import_user_row(self, row: dict):
        dn = row.get("distinguished_name", "").strip()
        sam = row.get("sam_account_name", "").strip()
        if not dn or not sam:
            raise ValueError(f"Missing required fields. sam_account_name='{sam}', dn='{dn}'")

        existing = self.db.query(User).filter(User.distinguished_name == dn).first()
        if existing:
            for key in ["sam_account_name", "display_name", "email", "department"]:
                if key in row and row[key]:
                    setattr(existing, key, row[key].strip())
        else:
            user = User(
                sam_account_name=sam,
                distinguished_name=dn,
                display_name=row.get("display_name", "").strip() or None,
                email=row.get("email", "").strip() or None,
                department=row.get("department", "").strip() or None,
            )
            self.db.add(user)

    def _import_group_row(self, row: dict):
        dn = row.get("distinguished_name", "").strip()
        name = row.get("name", "").strip()
        if not dn or not name:
            raise ValueError(f"Missing required fields for group. name='{name}', dn='{dn}'")

        existing = self.db.query(ADGroup).filter(ADGroup.distinguished_name == dn).first()
        if existing:
            for key in ["name", "display_name", "group_type", "group_scope", "description", "email"]:
                if key in row and row[key]:
                    setattr(existing, key, row[key].strip())
        else:
            group = ADGroup(
                name=name,
                distinguished_name=dn,
                display_name=row.get("display_name", "").strip() or None,
                group_type=row.get("group_type", "security").strip(),
                group_scope=row.get("group_scope", "global").strip(),
                description=row.get("description", "").strip() or None,
                email=row.get("email", "").strip() or None,
            )
            self.db.add(group)

    def _validate_columns(self, columns: list[str], entity_type: str):
        required = set(self.REQUIRED_COLUMNS.get(entity_type, []))
        present = set(columns)
        missing = required - present
        if missing:
            raise ValueError(
                f"Missing required columns for {entity_type}: {', '.join(sorted(missing))}"
            )
